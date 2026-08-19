"""B01Bill 사양: 지정 연월 매출로 거래처별 청구서 PDF를 만든다."""

from __future__ import annotations

import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import tkinter as tk
from collections.abc import Callable
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_DIALOG_DIR = Path("C:/")
DEFAULT_YEAR = "2025"

SALES_SHEET = "매출"
CUSTOMER_SHEET = "거래처"
TEMPLATE_SHEET = "청구서"
ORIGINAL_SHEETS = ("청구서", "청구서 BackUp", "청구서(sample)")

ITEM_START_ROW = 20
ITEM_DEFAULT_END_ROW = 35
DEFAULT_ITEM_ROWS = 16
INVALID_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]')
YEAR_PATTERN = re.compile(r"^\d{4}$")
CREATE_NO_WINDOW = 0x08000000
LogFn = Callable[[str], None]


def parse_year_month(year_text: str, month_text: str) -> tuple[str, str]:
    year = year_text.strip()
    if not YEAR_PATTERN.fullmatch(year):
        raise ValueError("연도는 4자리 숫자여야 합니다. 예: 2025")
    month_num = int(str(month_text).replace("월", "").strip())
    if month_num < 1 or month_num > 12:
        raise ValueError("월은 1부터 12까지여야 합니다.")
    month = f"{month_num:02d}"
    return f"{year}-{month}", f"{year[2:]}{month}"


def normalize_code(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1)[:-1].isdigit():
        return text[:-2]
    return text


def cell_text(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_sheet_frame(path: Path, sheet_name: str) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise FileNotFoundError(f"'{sheet_name}' 시트가 없습니다: {path}")
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return pd.DataFrame()
    header = [str(column).strip() if column is not None else "" for column in rows[0]]
    return pd.DataFrame(list(rows[1:]), columns=header)


def validate_files(sales_path: Path, bill_path: Path) -> None:
    if not sales_path.is_file():
        raise FileNotFoundError(f"엑셀 파일이 없습니다: {sales_path}")
    if not bill_path.is_file():
        raise FileNotFoundError(f"엑셀 파일이 없습니다: {bill_path}")
    sales_wb = load_workbook(sales_path, read_only=True)
    try:
        missing = [name for name in (SALES_SHEET, CUSTOMER_SHEET) if name not in sales_wb.sheetnames]
        if missing:
            raise FileNotFoundError(f"{sales_path}에 시트가 없습니다: {', '.join(missing)}")
    finally:
        sales_wb.close()
    bill_wb = load_workbook(bill_path, read_only=True)
    try:
        if TEMPLATE_SHEET not in bill_wb.sheetnames:
            raise FileNotFoundError(f"'{TEMPLATE_SHEET}' 시트가 없습니다: {bill_path}")
    finally:
        bill_wb.close()


def load_month_sales(sales_path: Path, year_month: str) -> pd.DataFrame:
    sales = read_sheet_frame(sales_path, SALES_SHEET)
    required = ["매출일", "상품명", "거래처코드", "매출수량", "단가"]
    missing = [name for name in required if name not in sales.columns]
    if missing:
        raise ValueError(f"매출 시트에 필요한 열이 없습니다: {', '.join(missing)}")
    sales = sales.copy()
    sales["거래처코드"] = sales["거래처코드"].map(normalize_code)
    sales["매출일"] = pd.to_datetime(sales["매출일"], errors="coerce")
    sales["매출수량"] = pd.to_numeric(sales["매출수량"], errors="coerce")
    sales["단가"] = pd.to_numeric(sales["단가"], errors="coerce")
    filtered = sales[sales["매출일"].dt.strftime("%Y-%m") == year_month]
    return filtered.sort_values(["거래처코드", "매출일"], kind="mergesort")


def load_customers(sales_path: Path) -> dict[str, dict[str, str]]:
    customers = read_sheet_frame(sales_path, CUSTOMER_SHEET)
    required = ["거래처코드", "거래처명", "우편번호", "주소"]
    missing = [name for name in required if name not in customers.columns]
    if missing:
        raise ValueError(f"거래처 시트에 필요한 열이 없습니다: {', '.join(missing)}")
    lookup: dict[str, dict[str, str]] = {}
    for _, row in customers.iterrows():
        code = normalize_code(row["거래처코드"])
        if not code or code in lookup:
            continue
        lookup[code] = {
            "거래처명": cell_text(row["거래처명"]),
            "우편번호": cell_text(row["우편번호"]),
            "주소": cell_text(row["주소"]),
        }
    return lookup


def unique_sheet_name(customer_name: str, yymm: str, used: set[str]) -> str:
    cleaned = INVALID_SHEET_CHARS.sub("_", customer_name).strip() or "거래처"
    index = 1
    while True:
        extra = f"_{index}" if index > 1 else ""
        suffix = f"_{yymm}{extra}"
        keep = 31 - len(suffix)
        if keep < 1:
            name = f"{yymm}{extra}"[:31]
        else:
            name = f"{cleaned[:keep]}{suffix}"
        if name not in used:
            used.add(name)
            return name
        index += 1


def fill_items(worksheet: Worksheet, items: pd.DataFrame) -> None:
    count = len(items)
    last_item_row = ITEM_START_ROW + count - 1
    if count > DEFAULT_ITEM_ROWS:
        extra = count - DEFAULT_ITEM_ROWS
        worksheet.insert_rows(ITEM_DEFAULT_END_ROW + 1, extra)
        tax_row = last_item_row + 1
        vat_row = tax_row + 1
        for row in range(ITEM_DEFAULT_END_ROW + 1, last_item_row + 1):
            worksheet[f"H{row}"] = f"=F{row}*G{row}"
        worksheet[f"H{tax_row}"] = f"=SUM(H{ITEM_START_ROW}:H{last_item_row})"
        worksheet[f"H{vat_row}"] = f"=INT(H{tax_row}*0.1)"
        worksheet["E17"] = f"=H{tax_row}"
        worksheet["F17"] = f"=H{tax_row}"
        worksheet["G17"] = f"=H{vat_row}"
        worksheet["H17"] = f"=H{vat_row}"
        worksheet["I17"] = f"=H{tax_row}+H{vat_row}"

    for offset, row in enumerate(items.itertuples(index=False)):
        excel_row = ITEM_START_ROW + offset
        sale_date = row.매출일
        worksheet[f"A{excel_row}"] = sale_date.strftime("%Y-%m-%d") if pd.notna(sale_date) else ""
        worksheet[f"B{excel_row}"] = "" if pd.isna(row.상품명) else str(row.상품명)
        worksheet[f"F{excel_row}"] = None if pd.isna(row.매출수량) else float(row.매출수량)
        worksheet[f"G{excel_row}"] = None if pd.isna(row.단가) else float(row.단가)


def restore_original_sheets(bill_path: Path) -> None:
    workbook = load_workbook(bill_path)
    for name in list(workbook.sheetnames):
        if name not in ORIGINAL_SHEETS:
            del workbook[name]
    workbook.save(bill_path)
    workbook.close()


def find_soffice() -> Path:
    candidates = [
        Path(os.environ.get("PROGRAMFILES", r"C:\Program Files")) / "LibreOffice" / "program" / "soffice.exe",
        Path(os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)")) / "LibreOffice" / "program" / "soffice.exe",
    ]
    path_cmd = shutil.which("soffice")
    if path_cmd:
        candidates.insert(0, Path(path_cmd))
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    raise FileNotFoundError("LibreOffice(soffice.exe)를 찾을 수 없습니다. LibreOffice를 설치하세요.")


def _convert_xlsx_batch(soffice: Path, xlsx_files: list[Path], out_dir: Path, profile_uri: str) -> None:
    command = [
        str(soffice),
        "--headless",
        "--norestore",
        "--nologo",
        "--nofirststartwizard",
        f"-env:UserInstallation={profile_uri}",
        "--convert-to",
        "pdf",
        "--outdir",
        str(out_dir),
    ]
    command.extend(str(path) for path in xlsx_files)
    kwargs: dict[str, object] = {"capture_output": True, "text": True, "check": False}
    if os.name == "nt":
        kwargs["creationflags"] = CREATE_NO_WINDOW
    completed = subprocess.run(command, **kwargs)
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout or "").strip()
        raise RuntimeError(f"LibreOffice PDF 변환 실패: {detail}")


def export_pdfs(bill_path: Path, pdf_dir: Path, sheet_names: list[str], log: LogFn | None = None) -> None:
    emit = log or (lambda _message: None)
    soffice = find_soffice()
    emit(f"LibreOffice 확인: {soffice}")
    pdf_dir.mkdir(parents=True, exist_ok=True)
    batch_size = 8
    total = len(sheet_names)
    with tempfile.TemporaryDirectory(prefix="bill_pdf_") as temp_dir:
        temp_path = Path(temp_dir)
        profile_dir = temp_path / "lo_profile"
        profile_dir.mkdir()
        profile_uri = "file:///" + str(profile_dir.resolve()).replace("\\", "/")
        xlsx_files: list[Path] = []
        for index, name in enumerate(sheet_names, start=1):
            emit(f"PDF 변환 준비 ({index}/{total}): {name}")
            single = load_workbook(bill_path)
            if name not in single.sheetnames:
                single.close()
                raise FileNotFoundError(f"내보낼 시트가 없습니다: {name}")
            for other in list(single.sheetnames):
                if other != name:
                    del single[other]
            xlsx_path = temp_path / f"{name}.xlsx"
            single.save(xlsx_path)
            single.close()
            xlsx_files.append(xlsx_path)

        batch_count = (len(xlsx_files) + batch_size - 1) // batch_size
        for batch_index, start in enumerate(range(0, len(xlsx_files), batch_size), start=1):
            batch = xlsx_files[start : start + batch_size]
            emit(f"PDF 변환 중 ({batch_index}/{batch_count}): {len(batch)}개")
            _convert_xlsx_batch(soffice, batch, temp_path, profile_uri)

        missing: list[str] = []
        for index, name in enumerate(sheet_names, start=1):
            produced = temp_path / f"{name}.pdf"
            if not produced.is_file():
                missing.append(name)
                continue
            target = pdf_dir / f"{name}.pdf"
            if target.exists():
                target.unlink()
            shutil.move(str(produced), str(target))
            emit(f"PDF 저장 ({index}/{total}): {target.name}")
        if missing:
            raise RuntimeError("PDF가 만들어지지 않은 시트: " + ", ".join(missing))


def create_bill_sheets(
    bill_path: Path,
    month_sales: pd.DataFrame,
    customers: dict[str, dict[str, str]],
    yymm: str,
    log: LogFn | None = None,
) -> tuple[list[str], list[str]]:
    emit = log or (lambda _message: None)
    skipped: list[str] = []
    created: list[str] = []
    used_names: set[str] = set(ORIGINAL_SHEETS)
    grouped = list(month_sales.groupby("거래처코드", sort=True))
    total = len(grouped)
    emit(f"청구서 시트 작성 시작: 거래처 {total}개")

    workbook = load_workbook(bill_path)
    try:
        if TEMPLATE_SHEET not in workbook.sheetnames:
            raise FileNotFoundError(f"'{TEMPLATE_SHEET}' 시트가 없습니다: {bill_path}")
        template = workbook[TEMPLATE_SHEET]
        for index, (code, items) in enumerate(grouped, start=1):
            master = customers.get(code)
            if master is None:
                skipped.append(code)
                emit(f"건너뜀 ({index}/{total}): 거래처코드 {code} (거래처 마스터 없음)")
                continue
            sheet_name = unique_sheet_name(master["거래처명"], yymm, used_names)
            copied = workbook.copy_worksheet(template)
            copied.title = sheet_name
            copied["B5"] = master["우편번호"]
            copied["B6"] = master["주소"]
            copied["B7"] = master["거래처명"]
            ordered = items.sort_values("매출일", kind="mergesort")[
                ["매출일", "상품명", "매출수량", "단가"]
            ]
            fill_items(copied, ordered)
            created.append(sheet_name)
            emit(f"시트 작성 ({index}/{total}): {sheet_name}")
        if created:
            emit("청구서.xlsx 저장 중...")
            workbook.save(bill_path)
            emit("청구서.xlsx 저장 완료")
    finally:
        workbook.close()
    return created, skipped


def format_summary(year_month: str, created: list[str], skipped: list[str], pdf_dir: Path) -> str:
    lines = [
        f"대상 연월: {year_month}",
        f"작성 시트: {len(created)}개",
    ]
    lines.extend(f"  - {name}" for name in created)
    lines.append(f"건너뛴 거래처: {len(skipped)}개")
    lines.extend(f"  - {code}" for code in skipped)
    lines.append(f"PDF 경로: {pdf_dir}")
    return "\n".join(lines)


def run_bills(
    sales_path: Path,
    bill_path: Path,
    pdf_dir: Path,
    year_text: str,
    month_text: str,
    log: LogFn | None = None,
) -> str:
    emit = log or (lambda _message: None)
    year_month, yymm = parse_year_month(year_text, month_text)
    emit(f"대상 연월: {year_month}")
    emit(f"매출 파일: {sales_path}")
    emit(f"청구서 파일: {bill_path}")
    emit(f"PDF 저장 폴더: {pdf_dir}")
    emit("엑셀 파일 확인 중...")
    validate_files(sales_path, bill_path)
    emit("매출 데이터 읽는 중...")
    month_sales = load_month_sales(sales_path, year_month)
    if month_sales.empty:
        raise ValueError(f"{year_month} 매출이 없습니다. 엑셀을 변경하지 않았습니다.")
    emit(f"대상 매출 {len(month_sales)}건")
    emit("거래처 마스터 읽는 중...")
    customers = load_customers(sales_path)
    emit(f"거래처 마스터 {len(customers)}건")
    created, skipped = create_bill_sheets(bill_path, month_sales, customers, yymm, log=emit)
    if not created:
        raise ValueError("작성할 청구서가 없습니다. 엑셀을 변경하지 않았습니다.\n" + format_summary(year_month, created, skipped, pdf_dir))
    emit(f"작성 시트 {len(created)}개, 건너뛴 거래처 {len(skipped)}개")
    try:
        emit("PDF 변환 시작...")
        export_pdfs(bill_path, pdf_dir, created, log=emit)
        emit("PDF 변환 완료")
    except Exception:
        emit("오류로 청구서 시트를 원상 복구합니다.")
        restore_original_sheets(bill_path)
        raise
    emit("청구서 시트를 원상 복구합니다.")
    restore_original_sheets(bill_path)
    emit("작업이 모두 끝났습니다. 창은 닫기 버튼으로 종료하세요.")
    return format_summary(year_month, created, skipped, pdf_dir)


class BillApp:
    def __init__(self, window: tk.Tk) -> None:
        self.window = window
        self.window.title("청구서 PDF 작성")
        self.window.minsize(760, 560)
        self.window.protocol("WM_DELETE_WINDOW", self.close)
        self.sales_path: Path | None = None
        self.bill_path: Path | None = None
        self.pdf_dir: Path | None = None
        self._running = False

        window.columnconfigure(0, weight=1)
        window.rowconfigure(0, weight=1)

        frame = ttk.Frame(window, padding=16)
        frame.grid(row=0, column=0, sticky="nsew")
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(5, weight=1)

        self.sales_label = tk.StringVar(value="선택되지 않음")
        self.bill_label = tk.StringVar(value="선택되지 않음")
        self.pdf_label = tk.StringVar(value="선택되지 않음")
        self.year_var = tk.StringVar(value=DEFAULT_YEAR)
        self.month_var = tk.StringVar(value="1")

        ttk.Button(frame, text="매출.xlsx 선택", command=self.pick_sales, width=22).grid(
            row=0, column=0, sticky="w", pady=4
        )
        ttk.Label(frame, textvariable=self.sales_label).grid(row=0, column=1, sticky="ew", padx=(12, 0))

        ttk.Button(frame, text="청구서.xlsx 선택", command=self.pick_bill, width=22).grid(
            row=1, column=0, sticky="w", pady=4
        )
        ttk.Label(frame, textvariable=self.bill_label).grid(row=1, column=1, sticky="ew", padx=(12, 0))

        year_month = ttk.Frame(frame)
        year_month.grid(row=2, column=0, columnspan=2, sticky="w", pady=8)
        ttk.Label(year_month, text="매출 연도").pack(side="left")
        ttk.Entry(year_month, textvariable=self.year_var, width=8).pack(side="left", padx=(8, 20))
        ttk.Label(year_month, text="매출 월").pack(side="left")
        month_box = ttk.Combobox(
            year_month,
            textvariable=self.month_var,
            values=[str(month) for month in range(1, 13)],
            width=6,
            state="readonly",
        )
        month_box.pack(side="left", padx=(8, 0))
        ttk.Label(year_month, text="월").pack(side="left", padx=(4, 0))

        ttk.Button(frame, text="PDF 저장 폴더 선택", command=self.pick_pdf_dir, width=22).grid(
            row=3, column=0, sticky="w", pady=4
        )
        ttk.Label(frame, textvariable=self.pdf_label).grid(row=3, column=1, sticky="ew", padx=(12, 0))

        ttk.Label(frame, text="실행 로그").grid(row=4, column=0, columnspan=2, sticky="w", pady=(12, 4))
        self.log_text = ScrolledText(frame, height=18, wrap="word", state="disabled")
        self.log_text.grid(row=5, column=0, columnspan=2, sticky="nsew")

        action_row = ttk.Frame(frame)
        action_row.grid(row=6, column=0, columnspan=2, sticky="e", pady=(16, 0))
        self.close_button = ttk.Button(action_row, text="닫기", command=self.close)
        self.close_button.pack(side="right")
        self.run_button = ttk.Button(action_row, text="실행", command=self.run)
        self.run_button.pack(side="right", padx=(0, 8))

        self.log("대기 중. 파일을 선택한 뒤 실행을 누르세요.")

    def _ask_open_excel(self, title: str) -> Path | None:
        chosen = filedialog.askopenfilename(
            parent=self.window,
            title=title,
            initialdir=str(DEFAULT_DIALOG_DIR),
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
        )
        return Path(chosen) if chosen else None

    def pick_sales(self) -> None:
        path = self._ask_open_excel("매출.xlsx 파일 선택")
        if path is not None:
            self.sales_path = path
            self.sales_label.set(str(path))

    def pick_bill(self) -> None:
        path = self._ask_open_excel("청구서.xlsx 파일 선택")
        if path is not None:
            self.bill_path = path
            self.bill_label.set(str(path))

    def pick_pdf_dir(self) -> None:
        chosen = filedialog.askdirectory(
            parent=self.window,
            title="PDF 저장 폴더 선택",
            initialdir=str(DEFAULT_DIALOG_DIR),
            mustexist=True,
        )
        if chosen:
            self.pdf_dir = Path(chosen)
            self.pdf_label.set(str(self.pdf_dir))

    def log(self, message: str) -> None:
        stamp = datetime.now().strftime("%H:%M:%S")
        line = f"[{stamp}] {message}\n"

        def append() -> None:
            self.log_text.configure(state="normal")
            self.log_text.insert(tk.END, line)
            self.log_text.see(tk.END)
            self.log_text.configure(state="disabled")

        if threading.current_thread() is threading.main_thread():
            append()
        else:
            self.window.after(0, append)

    def close(self) -> None:
        if self._running:
            self.log("작업이 진행 중입니다. 완료된 뒤 닫기 버튼을 누르세요.")
            return
        self.window.destroy()

    def _set_running(self, running: bool) -> None:
        self._running = running
        self.run_button.configure(state="disabled" if running else "normal")
        self.window.config(cursor="watch" if running else "")

    def run(self) -> None:
        if self._running:
            return
        if self.sales_path is None:
            messagebox.showwarning("청구서 작성", "매출.xlsx 파일을 선택하세요.", parent=self.window)
            return
        if self.bill_path is None:
            messagebox.showwarning("청구서 작성", "청구서.xlsx 파일을 선택하세요.", parent=self.window)
            return
        if self.pdf_dir is None:
            messagebox.showwarning("청구서 작성", "PDF 저장 폴더를 선택하세요.", parent=self.window)
            return
        self._set_running(True)
        self.log("실행을 시작합니다.")
        worker = threading.Thread(target=self._run_worker, daemon=True)
        worker.start()

    def _run_worker(self) -> None:
        try:
            summary = run_bills(
                self.sales_path,
                self.bill_path,
                self.pdf_dir,
                self.year_var.get(),
                self.month_var.get(),
                log=self.log,
            )
        except Exception as error:
            self.log(f"오류: {error}")
            self.window.after(0, lambda: self._finish(False))
        else:
            for line in summary.splitlines():
                self.log(line)
            self.window.after(0, lambda: self._finish(True))

    def _finish(self, success: bool) -> None:
        self._set_running(False)
        if success:
            self.log("실행이 종료되었습니다. 닫기 버튼으로 창을 닫으세요.")
        else:
            self.log("실행이 실패했습니다. 로그를 확인한 뒤 닫기 버튼으로 창을 닫으세요.")


def main() -> int:
    window = tk.Tk()
    BillApp(window)
    window.mainloop()
    return 0


if __name__ == "__main__":
    sys.exit(main())

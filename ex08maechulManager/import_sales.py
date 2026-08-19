"""csv 폴더의 CSV를 매출.xlsx 매출 시트에 이관한다."""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent
CSV_DIR = ROOT / "resData" / "csv"
COMPLETED_DIR = CSV_DIR / "completed"
XLSX_PATH = ROOT / "resData" / "매출.xlsx"
SHEET_NAME = "매출"

EXPECTED_COLUMNS = [
    "매출일",
    "영업소",
    "영업담당",
    "기간",
    "전표번호",
    "상품코드",
    "상품명",
    "대분류",
    "중분류",
    "소분류",
    "거래처코드",
    "거래처명",
    "매출수량",
    "매출금액",
    "매출이익",
    "단가",
]
NUMERIC_COLUMNS = ["매출수량", "매출금액", "매출이익", "단가"]
ENCODINGS = ("utf-8-sig", "utf-8", "cp949")


def collect_csv_files(csv_dir: Path) -> list[Path]:
    if not csv_dir.is_dir():
        return []
    files = [path for path in csv_dir.iterdir() if path.is_file() and path.suffix.lower() == ".csv"]
    return sorted(files, key=lambda path: path.name)


def read_csv_file(path: Path) -> pd.DataFrame:
    last_error: Exception | None = None
    for encoding in ENCODINGS:
        try:
            frame = pd.read_csv(path, encoding=encoding, dtype=str, keep_default_na=False)
            break
        except (UnicodeDecodeError, UnicodeError, OSError) as error:
            last_error = error
    else:
        raise RuntimeError(f"인코딩을 읽을 수 없습니다: {path.name}") from last_error

    columns = [str(column).strip() for column in frame.columns]
    if columns != EXPECTED_COLUMNS:
        raise ValueError(f"헤더가 지정 컬럼과 다릅니다: {path.name} -> {columns}")

    frame.columns = EXPECTED_COLUMNS
    return frame


def convert_numeric_columns(frame: pd.DataFrame) -> pd.DataFrame:
    converted = frame.copy()
    for column in NUMERIC_COLUMNS:
        converted[column] = pd.to_numeric(converted[column], errors="coerce")
    return converted


def validate_workbook(xlsx_path: Path) -> None:
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"엑셀 파일이 없습니다: {xlsx_path}")
    workbook = load_workbook(xlsx_path, read_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise FileNotFoundError(f"'{SHEET_NAME}' 시트가 없습니다: {xlsx_path}")
    finally:
        workbook.close()


def clear_sheet(worksheet: Worksheet) -> None:
    merged_ranges = list(worksheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        worksheet.unmerge_cells(str(merged_range))
    if worksheet.max_row and worksheet.max_row >= 1:
        worksheet.delete_rows(1, worksheet.max_row)


def write_sales_sheet(xlsx_path: Path, frame: pd.DataFrame) -> None:
    workbook = load_workbook(xlsx_path)
    if SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise FileNotFoundError(f"'{SHEET_NAME}' 시트가 없습니다: {xlsx_path}")

    worksheet = workbook[SHEET_NAME]
    clear_sheet(worksheet)
    for row in dataframe_to_rows(frame, index=False, header=True):
        worksheet.append(row)

    workbook.save(xlsx_path)
    workbook.close()


def move_completed_files(success_files: list[Path], completed_dir: Path) -> None:
    completed_dir.mkdir(parents=True, exist_ok=True)
    for source in success_files:
        destination = completed_dir / source.name
        if destination.exists():
            destination.unlink()
        shutil.move(str(source), str(destination))


def print_summary(
    *,
    collected_count: int,
    success_files: list[Path],
    failed_files: list[tuple[str, str]],
    row_count: int,
    moved: bool,
) -> None:
    print(f"수집 CSV: {collected_count}개")
    print(f"성공: {len(success_files)}개")
    for path in success_files:
        print(f"  - {path.name}")
    print(f"실패: {len(failed_files)}개")
    for name, reason in failed_files:
        print(f"  - {name}: {reason}")
    print(f"기록 행 수: {row_count}")
    print(f"엑셀: {XLSX_PATH}")
    print(f"시트: {SHEET_NAME}")
    if moved:
        print(f"이동 경로: {COMPLETED_DIR}")
    else:
        print("파일 이동: 없음")


def main() -> int:
    csv_files = collect_csv_files(CSV_DIR)
    if not csv_files:
        print(f"CSV 파일이 없습니다: {CSV_DIR}")
        print("엑셀을 변경하지 않고 종료합니다.")
        return 0

    try:
        validate_workbook(XLSX_PATH)
    except FileNotFoundError as error:
        print(f"오류: {error}")
        print("엑셀을 변경하지 않았고, 파일도 이동하지 않았습니다.")
        return 1

    success_frames: list[pd.DataFrame] = []
    success_files: list[Path] = []
    failed_files: list[tuple[str, str]] = []

    for path in csv_files:
        try:
            frame = read_csv_file(path)
        except Exception as error:  # noqa: BLE001 - 파일 단위로 건너뛰기
            failed_files.append((path.name, str(error)))
            continue
        success_frames.append(frame)
        success_files.append(path)

    if not success_files:
        print_summary(
            collected_count=len(csv_files),
            success_files=success_files,
            failed_files=failed_files,
            row_count=0,
            moved=False,
        )
        print("성공한 CSV가 없어 엑셀을 변경하지 않았습니다.")
        return 1

    merged = convert_numeric_columns(pd.concat(success_frames, ignore_index=True))

    try:
        write_sales_sheet(XLSX_PATH, merged)
    except Exception as error:  # noqa: BLE001 - 저장 실패 시 이동 금지
        print(f"엑셀 저장 실패: {error}")
        print("파일 이동을 하지 않았습니다.")
        print_summary(
            collected_count=len(csv_files),
            success_files=success_files,
            failed_files=failed_files,
            row_count=len(merged),
            moved=False,
        )
        return 1

    move_completed_files(success_files, COMPLETED_DIR)
    print_summary(
        collected_count=len(csv_files),
        success_files=success_files,
        failed_files=failed_files,
        row_count=len(merged),
        moved=True,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())

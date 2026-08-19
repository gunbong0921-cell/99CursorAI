"""A01CSV 사양: csv 폴더의 CSV를 매출.xlsx 매출 시트에 이관한다."""

from __future__ import annotations

import shutil
import sys
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet._reader import WorksheetReader
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_DIALOG_DIR = Path("C:/")
SHEET_NAME = "매출"

EXPECTED_COLUMNS = [
    "매출일",
    "영업소",
    "영업담당",
    "기간",
    "전표번호",
    "상품코드",
    "상품명",
    "대분류",
    "중분류",
    "소분류",
    "거래처코드",
    "거래처명",
    "매출수량",
    "매출금액",
    "매출이익",
    "단가",
]
NUMERIC_COLUMNS = ["매출수량", "매출금액", "매출이익", "단가"]
ENCODINGS = ("utf-8-sig", "utf-8", "cp949")


def select_csv_and_excel() -> tuple[Path, Path] | None:
    """CSV 원본 폴더와 매출.xlsx를 대화 상자로 고른다. 기본 경로는 C:\\ 이다."""
    root = tk.Tk()
    root.withdraw()
    root.wm_attributes("-topmost", True)
    initial_dir = str(DEFAULT_DIALOG_DIR)
    csv_dir = filedialog.askdirectory(
        parent=root,
        title="CSV 원본 폴더 선택",
        initialdir=initial_dir,
        mustexist=True,
    )
    if not csv_dir:
        root.destroy()
        return None
    xlsx_path = filedialog.askopenfilename(
        parent=root,
        title="매출.xlsx 파일 선택",
        initialdir=initial_dir,
        filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
    )
    root.destroy()
    if not xlsx_path:
        return None
    return Path(csv_dir), Path(xlsx_path)


def collect_csv_files(csv_dir: Path) -> list[Path]:
    """csv 폴더 바로 아래의 .csv만 수집한다. completed 등 하위 폴더는 제외한다."""
    if not csv_dir.is_dir():
        return []
    files = [
        path
        for path in csv_dir.iterdir()
        if path.is_file() and path.suffix.lower() == ".csv"
    ]
    return sorted(files, key=lambda path: path.name)


def read_csv_file(path: Path) -> pd.DataFrame:
    """UTF-8을 우선 시도하고, 실패하면 CP949로 읽는다. 헤더가 지정 컬럼과 같아야 한다."""
    last_error: Exception | None = None
    frame: pd.DataFrame | None = None
    for encoding in ENCODINGS:
        try:
            frame = pd.read_csv(path, encoding=encoding, dtype=str, keep_default_na=False)
            break
        except (UnicodeDecodeError, UnicodeError, OSError) as error:
            last_error = error

    if frame is None:
        raise RuntimeError(f"인코딩을 읽을 수 없습니다: {path.name}") from last_error

    columns = [str(column).strip() for column in frame.columns]
    if columns != EXPECTED_COLUMNS:
        raise ValueError(f"헤더가 지정 컬럼과 다릅니다: {path.name} -> {columns}")

    frame.columns = EXPECTED_COLUMNS
    return frame


def convert_numeric_columns(frame: pd.DataFrame) -> pd.DataFrame:
    converted = frame.copy()
    for column in NUMERIC_COLUMNS:
        converted[column] = pd.to_numeric(converted[column], errors="coerce")
    return converted


def _bind_col_dimensions_safe(self) -> None:
    styles = self.ws.parent._cell_styles
    for col, cd in self.parser.column_dimensions.items():
        if "style" in cd:
            key = int(cd["style"])
            if 0 <= key < len(styles):
                cd["style"] = styles[key]
            else:
                del cd["style"]
        self.ws.column_dimensions[col] = ColumnDimension(self.ws, **cd)


def _bind_row_dimensions_safe(self) -> None:
    styles = self.ws.parent._cell_styles
    for row, rd in self.parser.row_dimensions.items():
        if "s" in rd:
            key = int(rd["s"])
            if 0 <= key < len(styles):
                rd["s"] = styles[key]
            else:
                del rd["s"]
        self.ws.row_dimensions[int(row)] = RowDimension(self.ws, **rd)


def _bind_cells_safe(self) -> None:
    styles = self.ws.parent._cell_styles
    fallback = styles[0] if styles else None
    for _idx, row in self.parser.parse():
        for cell in row:
            style_id = cell["style_id"]
            style = styles[style_id] if 0 <= style_id < len(styles) else fallback
            created = Cell(self.ws, row=cell["row"], column=cell["column"], style_array=style)
            created._value = cell["value"]
            created.data_type = cell["data_type"]
            self.ws._cells[(cell["row"], cell["column"])] = created
    if self.ws._cells:
        self.ws._current_row = self.ws.max_row


def _bind_formatting_safe(self) -> None:
    differentials = self.ws.parent._differential_styles
    for cf in self.parser.formatting:
        for rule in cf.rules:
            if rule.dxfId is not None:
                if 0 <= rule.dxfId < len(differentials):
                    rule.dxf = differentials[rule.dxfId]
                else:
                    rule.dxfId = None
            self.ws.conditional_formatting[cf] = rule


WorksheetReader.bind_cells = _bind_cells_safe
WorksheetReader.bind_formatting = _bind_formatting_safe
WorksheetReader.bind_col_dimensions = _bind_col_dimensions_safe
WorksheetReader.bind_row_dimensions = _bind_row_dimensions_safe


def validate_workbook(xlsx_path: Path) -> None:
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"엑셀 파일이 없습니다: {xlsx_path}")
    workbook = load_workbook(xlsx_path, read_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise FileNotFoundError(f"'{SHEET_NAME}' 시트가 없습니다: {xlsx_path}")
    finally:
        workbook.close()


def clear_sheet(worksheet: Worksheet) -> None:
    merged_ranges = list(worksheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        worksheet.unmerge_cells(str(merged_range))
    if worksheet.max_row and worksheet.max_row >= 1:
        worksheet.delete_rows(1, worksheet.max_row)


def write_sales_sheet(xlsx_path: Path, frame: pd.DataFrame) -> None:
    workbook = load_workbook(xlsx_path)
    if SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise FileNotFoundError(f"'{SHEET_NAME}' 시트가 없습니다: {xlsx_path}")

    worksheet = workbook[SHEET_NAME]
    clear_sheet(worksheet)
    for row in dataframe_to_rows(frame, index=False, header=True):
        worksheet.append(row)

    workbook.save(xlsx_path)
    workbook.close()


def move_completed_files(success_files: list[Path], completed_dir: Path) -> None:
    completed_dir.mkdir(parents=True, exist_ok=True)
    for source in success_files:
        destination = completed_dir / source.name
        if destination.exists():
            destination.unlink()
        shutil.move(str(source), str(destination))


def notify(title: str, message: str, kind: str = "info") -> None:
    print(message)
    popup = tk.Tk()
    popup.withdraw()
    popup.wm_attributes("-topmost", True)
    if kind == "error":
        messagebox.showerror(title, message, parent=popup)
    elif kind == "warning":
        messagebox.showwarning(title, message, parent=popup)
    else:
        messagebox.showinfo(title, message, parent=popup)
    popup.destroy()


def format_summary(
    *,
    collected_count: int,
    success_files: list[Path],
    failed_files: list[tuple[str, str]],
    row_count: int,
    moved: bool,
    xlsx_path: Path,
    completed_dir: Path,
) -> str:
    lines = [
        f"수집 CSV: {collected_count}개",
        f"성공: {len(success_files)}개",
    ]
    lines.extend(f"  - {path.name}" for path in success_files)
    lines.append(f"실패: {len(failed_files)}개")
    lines.extend(f"  - {name}: {reason}" for name, reason in failed_files)
    lines.extend(
        [
            f"기록 행 수: {row_count}",
            f"엑셀: {xlsx_path}",
            f"시트: {SHEET_NAME}",
            f"이동 경로: {completed_dir}" if moved else "파일 이동: 없음",
        ]
    )
    return "\n".join(lines)


def print_summary(
    *,
    collected_count: int,
    success_files: list[Path],
    failed_files: list[tuple[str, str]],
    row_count: int,
    moved: bool,
    xlsx_path: Path,
    completed_dir: Path,
) -> None:
    print(
        format_summary(
            collected_count=collected_count,
            success_files=success_files,
            failed_files=failed_files,
            row_count=row_count,
            moved=moved,
            xlsx_path=xlsx_path,
            completed_dir=completed_dir,
        )
    )


def main() -> int:
    selected = select_csv_and_excel()
    if selected is None:
        return 0

    csv_dir, xlsx_path = selected
    completed_dir = csv_dir / "completed"
    print(f"CSV 원본 폴더: {csv_dir}")
    print(f"매출 엑셀: {xlsx_path}")

    csv_files = collect_csv_files(csv_dir)
    if not csv_files:
        notify("CSV 이관", f"CSV 파일이 없습니다.\n{csv_dir}\n엑셀을 변경하지 않고 종료합니다.", "warning")
        return 0

    try:
        validate_workbook(xlsx_path)
    except FileNotFoundError as error:
        notify("CSV 이관", f"오류: {error}\n엑셀을 변경하지 않았고, 파일도 이동하지 않았습니다.", "error")
        return 1

    success_frames: list[pd.DataFrame] = []
    success_files: list[Path] = []
    failed_files: list[tuple[str, str]] = []

    for path in csv_files:
        try:
            frame = read_csv_file(path)
        except (RuntimeError, ValueError, OSError, pd.errors.ParserError) as error:
            failed_files.append((path.name, str(error)))
            continue
        success_frames.append(frame)
        success_files.append(path)

    if not success_files:
        summary = format_summary(
            collected_count=len(csv_files),
            success_files=success_files,
            failed_files=failed_files,
            row_count=0,
            moved=False,
            xlsx_path=xlsx_path,
            completed_dir=completed_dir,
        )
        print_summary(
            collected_count=len(csv_files),
            success_files=success_files,
            failed_files=failed_files,
            row_count=0,
            moved=False,
            xlsx_path=xlsx_path,
            completed_dir=completed_dir,
        )
        notify("CSV 이관", summary + "\n성공한 CSV가 없어 엑셀을 변경하지 않았습니다.", "error")
        return 1

    merged = convert_numeric_columns(pd.concat(success_frames, ignore_index=True))

    try:
        write_sales_sheet(xlsx_path, merged)
    except (OSError, IndexError, ValueError, KeyError) as error:
        summary = format_summary(
            collected_count=len(csv_files),
            success_files=success_files,
            failed_files=failed_files,
            row_count=len(merged),
            moved=False,
            xlsx_path=xlsx_path,
            completed_dir=completed_dir,
        )
        print_summary(
            collected_count=len(csv_files),
            success_files=success_files,
            failed_files=failed_files,
            row_count=len(merged),
            moved=False,
            xlsx_path=xlsx_path,
            completed_dir=completed_dir,
        )
        notify("CSV 이관", f"엑셀 저장 실패: {error}\n파일 이동을 하지 않았습니다.\n\n{summary}", "error")
        return 1

    move_completed_files(success_files, completed_dir)
    summary = format_summary(
        collected_count=len(csv_files),
        success_files=success_files,
        failed_files=failed_files,
        row_count=len(merged),
        moved=True,
        xlsx_path=xlsx_path,
        completed_dir=completed_dir,
    )
    print_summary(
        collected_count=len(csv_files),
        success_files=success_files,
        failed_files=failed_files,
        row_count=len(merged),
        moved=True,
        xlsx_path=xlsx_path,
        completed_dir=completed_dir,
    )
    notify("CSV 이관", summary, "info")
    return 0


if __name__ == "__main__":
    sys.exit(main())
